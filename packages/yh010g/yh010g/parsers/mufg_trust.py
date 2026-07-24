"""三菱UFJ信託銀行 会社別議案別行使結果パーサ (2017年〜の全形式対応)。

形式の変遷 (2026-07-23/24 実査):
  新形式 (2023年頃〜): シート「議決権行使結果」(末尾空白の揺れあり)。列:
    銘柄コード / 銘柄名称 / 総会日(YYYYMMDD文字列) / 総会種類(定時総会) / 議案番号 /
    子議案番号 / 提案者(会社|株主) / 議案分類 / 賛否 / 理由
  旧形式 (2017〜2022年頃): シート「議決権行使結果」または「会社別議案別行使結果」。列:
    証券コード / 社名 / 総会日(datetime) / 総会種類(定時) / 提案者 / 議案番号('01') /
    議案分類 / 候補者番号 / 賛否 / 理由
列順・列名が異なるため、ヘッダ行の列名からインデックスを構築する。
賛否の「*」(顧客ガイドラインによる不統一行使の注記) は値正規化側で除去。
"""

from __future__ import annotations

import unicodedata
import warnings

import openpyxl

from yh010g.schema import (
    UnifiedRecord, map_proposer, map_vote, normalize_date, normalize_sec_code, normalize_sub_no,
)

MANAGER = "mufg_trust"
SHEET_NAMES = {"議決権行使結果", "会社別議案別行使結果"}
COLUMN_ALIASES = {
    "sec_code": {"銘柄コード", "証券コード"},
    "company_name": {"銘柄名称", "社名"},
    "meeting_date": {"総会日"},
    "meeting_type": {"総会種類"},
    "proposal_no": {"議案番号"},
    "sub_no": {"子議案番号", "候補者番号"},
    "proposer": {"提案者"},
    "category": {"議案分類"},
    "vote": {"賛否"},
    "reason": {"理由"},
}


def _norm(s) -> str:
    return unicodedata.normalize("NFKC", str(s)).strip()


def _map_header(cells: list) -> dict[str, int] | None:
    names = [_norm(c) for c in cells]
    idx: dict[str, int] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for j, n in enumerate(names):
            # 「理由　（詳細は…」のような注記付きヘッダは前方一致で拾う
            if n in aliases or any(n.startswith(a) for a in aliases):
                idx[field] = j
                break
    required = set(COLUMN_ALIASES) - {"reason", "sub_no"}
    return idx if required <= idx.keys() else None


def parse_mufg_trust(path: str) -> list[UnifiedRecord]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        matches = [n for n in wb.sheetnames if n.strip() in SHEET_NAMES]
        if not matches:
            raise ValueError(f"{path}: no known sheet (got {wb.sheetnames})")
        ws = wb[matches[0]]
        idx: dict[str, int] | None = None
        out: list[UnifiedRecord] = []
        unknown_votes: list[str] = []
        skipped_footers: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [c if c is not None else "" for c in row]
            if idx is None:
                idx = _map_header(cells)
                continue
            if _norm(cells[idx["sec_code"]]) == "":
                continue
            try:
                sec_code = normalize_sec_code(cells[idx["sec_code"]])
            except ValueError:
                skipped_footers.append(_norm(cells[idx["sec_code"]])[:20])
                continue
            vote_raw = cells[idx["vote"]]
            vote, known = map_vote(vote_raw)
            if not known:
                unknown_votes.append(str(vote_raw))
            out.append(UnifiedRecord(
                manager=MANAGER,
                sec_code=sec_code,
                company_name=_norm(cells[idx["company_name"]]),
                meeting_date=normalize_date(cells[idx["meeting_date"]]),
                meeting_type=_norm(cells[idx["meeting_type"]]),
                proposal_no=int(float(_norm(cells[idx["proposal_no"]]))),
                sub_no=normalize_sub_no(cells[idx["sub_no"]]) if "sub_no" in idx else 0,
                proposer=map_proposer(cells[idx["proposer"]]),
                category=_norm(cells[idx["category"]]),
                vote=vote,
                vote_raw=_norm(vote_raw),
                reason=_norm(cells[idx["reason"]]) if "reason" in idx else "",
            ))
        if idx is None:
            raise ValueError(f"{path}: header row not found")
        if not out:
            raise ValueError(f"{path}: no data rows parsed")
        if unknown_votes:
            warnings.warn(f"{path}: {len(unknown_votes)} unknown vote values, e.g. {unknown_votes[:5]}",
                          stacklevel=2)
        return out
    finally:
        wb.close()
