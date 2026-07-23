"""アモーヴァ・アセットマネジメント (旧日興AM) 個別開示パーサ。

形式 (2026-07-23 実査):
  シートは 24q1 まで 'JP' 単独、25q1 以降 '集計結果' + '個別開示'。
  個別開示のヘッダ (4行目):
  企業コード / 企業名 / 総会種類 / 総会日 / 親議案番号 / 子議案番号 / 議案分類 /
  提案者(会社提案|株主提案) / 弊社賛否 / 賛否理由 / 備考
  総会日は datetime か 'YYYYMMDD' 文字列の両方がありうる (期により混在)。
"""

from __future__ import annotations

import unicodedata
import warnings

import openpyxl

from yh010g.schema import (
    UnifiedRecord, map_proposer, map_vote, normalize_date, normalize_sec_code, normalize_sub_no,
)

HEADER_PREFIX = ["企業コード", "企業名", "総会種類", "総会日", "親議案番号", "子議案番号", "議案分類", "提案者", "弊社賛否"]
MANAGER = "amova"


def _pick_sheet(wb) -> str:
    for name in ("個別開示", "JP"):
        if name in wb.sheetnames:
            return name
    raise ValueError(f"no known sheet in {wb.sheetnames}")


def parse_amova(path: str) -> list[UnifiedRecord]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb[_pick_sheet(wb)]
        header_seen = False
        out: list[UnifiedRecord] = []
        unknown_votes: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [c if c is not None else "" for c in row]
            if not header_seen:
                head = [unicodedata.normalize("NFKC", str(c)).strip() for c in cells[:9]]
                if head == HEADER_PREFIX:
                    header_seen = True
                continue
            if str(cells[0]).strip() == "":
                continue
            vote, known = map_vote(cells[8])
            if not known:
                unknown_votes.append(str(cells[8]))
            out.append(UnifiedRecord(
                manager=MANAGER,
                sec_code=normalize_sec_code(cells[0]),
                company_name=str(cells[1]).strip(),
                meeting_date=normalize_date(cells[3]),
                meeting_type=str(cells[2]).strip(),
                proposal_no=int(float(str(cells[4]))),
                sub_no=normalize_sub_no(cells[5]),
                proposer=map_proposer(cells[7]),
                category=str(cells[6]).strip(),
                vote=vote,
                vote_raw=str(cells[8]).strip(),
                reason=str(cells[9]).strip() if len(cells) > 9 else "",
            ))
        if not header_seen:
            raise ValueError(f"{path}: header row not found")
        if not out:
            raise ValueError(f"{path}: no data rows parsed")
        if unknown_votes:
            warnings.warn(f"{path}: {len(unknown_votes)} unknown vote values, e.g. {unknown_votes[:5]}",
                          stacklevel=2)
        return out
    finally:
        wb.close()
