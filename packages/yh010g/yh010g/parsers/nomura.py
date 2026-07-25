"""野村アセットマネジメント 個別開示パーサ。

形式 (2026-07-23 実査): 単一シート (シート名は社名)。ヘッダ行は「企業\nコード」で検出
(2024年版は3行目・2025年版は4行目)。列:
  企業コード / 企業名 / 総会種類(定時) / 総会日(YYYY/MM/DD) / 提案者(会社|株主) /
  議案番号('1' or '2.1' 文字列) / 議案分類 / 賛否 / 理由
賛否は当社基準による値 (顧客基準による差異は開示対象外と注記あり)。
"""

from __future__ import annotations

import unicodedata
import warnings

import openpyxl

from yh010g.schema import (
    UnifiedRecord, map_proposer, map_vote, normalize_date, normalize_sec_code, parse_proposal_no,
)

MANAGER = "nomura"


def parse_nomura(path: str) -> list[UnifiedRecord]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_seen = False
        out: list[UnifiedRecord] = []
        unknown_votes: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [c if c is not None else "" for c in row]
            first = unicodedata.normalize("NFKC", str(cells[0])).replace("\n", "").strip()
            if not header_seen:
                if first == "企業コード":
                    header_seen = True
                continue
            if str(cells[0]).strip() == "":
                continue
            pno, sub = parse_proposal_no(cells[5])
            vote, known = map_vote(cells[7])
            if not known:
                unknown_votes.append(str(cells[7]))
            out.append(UnifiedRecord(
                manager=MANAGER,
                sec_code=normalize_sec_code(cells[0]),
                company_name=str(cells[1]).strip(),
                meeting_date=normalize_date(cells[3]),
                meeting_type=str(cells[2]).strip(),
                proposal_no=pno,
                sub_no=sub,
                proposer=map_proposer(cells[4]),
                category=str(cells[6]).strip(),
                vote=vote,
                vote_raw=str(cells[7]).strip(),
                reason=str(cells[8]).strip() if len(cells) > 8 else "",
            ))
        if not header_seen:
            raise ValueError(f"{path}: header row not found")
        if not out:
            raise ValueError(f"{path}: no data rows parsed")
        if unknown_votes:
            warnings.warn(f"{path}: {len(unknown_votes)} unknown vote values, e.g. {unknown_votes[:5]}",
                          stacklevel=2)
        return out
    finally:
        wb.close()
