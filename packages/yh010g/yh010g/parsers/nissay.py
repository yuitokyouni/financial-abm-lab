"""ニッセイアセットマネジメント 個別開示パーサ (会社ブロック形式)。

形式 (2026-07-23 実査):
  シート「議案別行使結果（4・5月）」「議案別行使結果（6月）」等 (REIT シートは除外可能)。
  会社ブロック:
    行1: '（株）極洋(1301)'  ← 社名 + (証券コード)
    行2: '株主総会開催日　2025年06月25日'
    行3: ヘッダ '議案\n番号 / 候補者\n番号 / 議案区分 / 提案 / 判断 / 主な判断理由'
    行4〜: データ行 (空行でブロック終端)
  総会種類は開示されない → meeting_type = ""。
"""

from __future__ import annotations

import re
import unicodedata
import warnings

import openpyxl

from yh010g.schema import (
    UnifiedRecord, map_proposer, map_vote, normalize_date, normalize_sec_code, normalize_sub_no,
)

MANAGER = "nissay"
SHEET_PAT = re.compile(r"^議案別行使結果")
COMPANY_PAT = re.compile(r"^(?P<name>.+?)[(（](?P<code>[0-9A-Z]{4,5})[)）]\s*$")
DATE_PAT = re.compile(r"株主総会開催日\s*(?P<date>\d{4}年\d{1,2}月\d{1,2}日)")


def _norm(s) -> str:
    return unicodedata.normalize("NFKC", str(s)).strip()


def parse_nissay(path: str, include_reit: bool = False) -> list[UnifiedRecord]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        sheets = [n for n in wb.sheetnames if SHEET_PAT.match(n)]
        if not include_reit:
            sheets = [n for n in sheets if "REIT" not in n.upper()]
        if not sheets:
            raise ValueError(f"{path}: no 議案別行使結果 sheets (got {wb.sheetnames})")
        out: list[UnifiedRecord] = []
        unknown_votes: list[str] = []
        for sn in sheets:
            ws = wb[sn]
            company, code, mdate = None, None, None
            in_block = False
            for row in ws.iter_rows(values_only=True):
                cells = [c if c is not None else "" for c in row]
                c0 = _norm(cells[0])
                m = COMPANY_PAT.match(c0)
                if m and not DATE_PAT.search(c0):
                    company, code = m.group("name").strip(), m.group("code")
                    mdate, in_block = None, False
                    continue
                dm = DATE_PAT.search(c0)
                if dm:
                    mdate = normalize_date(dm.group("date"))
                    continue
                if c0.startswith("議案"):  # ヘッダ行 ('議案\n番号')
                    in_block = company is not None and mdate is not None
                    continue
                if not in_block:
                    continue
                if c0 == "":
                    in_block = False
                    continue
                vote, known = map_vote(cells[4])
                if not known:
                    unknown_votes.append(str(cells[4]))
                out.append(UnifiedRecord(
                    manager=MANAGER,
                    sec_code=normalize_sec_code(code),
                    company_name=company,
                    meeting_date=mdate,
                    meeting_type="",
                    proposal_no=int(float(c0)),
                    sub_no=normalize_sub_no(cells[1]),
                    proposer=map_proposer(cells[3]),
                    category=_norm(cells[2]),
                    vote=vote,
                    vote_raw=_norm(cells[4]),
                    reason=_norm(cells[5]) if len(cells) > 5 else "",
                ))
        if not out:
            raise ValueError(f"{path}: no data rows parsed")
        if unknown_votes:
            warnings.warn(f"{path}: {len(unknown_votes)} unknown vote values, e.g. {unknown_votes[:5]}",
                          stacklevel=2)
        return out
    finally:
        wb.close()
