"""三井住友DSアセットマネジメント 個別開示パーサ。

形式 (2026-07-23 実査): 単一シート (名称は 'smdam_votingresults' / 'smdam_votingresult_Apr-Jun-2024'
等の揺れあり → 先頭シートを使用)。1行目がヘッダ:
  銘柄コード / 会社名 / 総会区分(定時総会) / 総会日(YYYYMMDD) / 議案番号('1'|'2-1') /
  提案者(会社提案|株主提案) / 賛成/反対 / 議案分類 / 主な判断理由
"""

from __future__ import annotations

import unicodedata
import warnings

import openpyxl

from yh010g.schema import (
    UnifiedRecord, map_proposer, map_vote, normalize_date, normalize_sec_code, parse_proposal_no,
)

MANAGER = "smdam"


def parse_smdam(path: str) -> list[UnifiedRecord]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_seen = False
        out: list[UnifiedRecord] = []
        unknown_votes: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [c if c is not None else "" for c in row]
            first = unicodedata.normalize("NFKC", str(cells[0])).strip()
            if not header_seen:
                if first == "銘柄コード":
                    header_seen = True
                continue
            if first == "":
                continue
            pno, sub = parse_proposal_no(cells[4])
            vote, known = map_vote(cells[6])
            if not known:
                unknown_votes.append(str(cells[6]))
            out.append(UnifiedRecord(
                manager=MANAGER,
                sec_code=normalize_sec_code(cells[0]),
                company_name=str(cells[1]).strip(),
                meeting_date=normalize_date(cells[3]),
                meeting_type=str(cells[2]).strip(),
                proposal_no=pno,
                sub_no=sub,
                proposer=map_proposer(cells[5]),
                category=str(cells[7]).strip(),
                vote=vote,
                vote_raw=str(cells[6]).strip(),
                reason=str(cells[8]).strip() if len(cells) > 8 else "",
            ))
        if not header_seen:
            raise ValueError(f"{path}: header row not found")
        if not out:
            raise ValueError(f"{path}: no data rows parsed")
        if unknown_votes:
            warnings.warn(f"{path}: {len(unknown_votes)} unknown vote values, e.g. {unknown_votes[:5]}",
                          stacklevel=2)
        return out
    finally:
        wb.close()
