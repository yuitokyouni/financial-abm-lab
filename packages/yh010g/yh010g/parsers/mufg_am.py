"""三菱UFJアセットマネジメント 個別開示パーサ (信託銀行とは別会社・別形式)。

形式 (2026-07-23 実査): シート「個別議案行使結果」(2024年版には空の「議決権行使結果」
シートが先行するため名前で選択)。1行目がヘッダ:
  銘柄コード / 銘柄名称 / 総会日(YYYYMMDD) / 総会種類(定時) / 提案者(会社|株主) /
  議案番号 / 子議案番号 / 議案分類 / 判断 / 理由
"""

from __future__ import annotations

import unicodedata
import warnings

import openpyxl

from yh010g.schema import (
    UnifiedRecord, map_proposer, map_vote, normalize_date, normalize_sec_code, normalize_sub_no,
)

MANAGER = "mufg_am"
SHEET = "個別議案行使結果"


def parse_mufg_am(path: str) -> list[UnifiedRecord]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        if SHEET not in wb.sheetnames:
            raise ValueError(f"{path}: sheet {SHEET!r} not found (got {wb.sheetnames})")
        ws = wb[SHEET]
        header_seen = False
        out: list[UnifiedRecord] = []
        unknown_votes: list[str] = []
        skipped: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [c if c is not None else "" for c in row]
            first = unicodedata.normalize("NFKC", str(cells[0])).strip()
            if not header_seen:
                if first == "銘柄コード":
                    header_seen = True
                continue
            if first == "":
                continue
            try:
                sec_code = normalize_sec_code(cells[0])
            except ValueError:
                skipped.append(first[:20])
                continue
            vote, known = map_vote(cells[8])
            if not known:
                unknown_votes.append(str(cells[8]))
            out.append(UnifiedRecord(
                manager=MANAGER,
                sec_code=sec_code,
                company_name=str(cells[1]).strip(),
                meeting_date=normalize_date(cells[2]),
                meeting_type=str(cells[3]).strip(),
                proposal_no=int(float(str(cells[5]))),
                sub_no=normalize_sub_no(cells[6]),
                proposer=map_proposer(cells[4]),
                category=str(cells[7]).strip(),
                vote=vote,
                vote_raw=str(cells[8]).strip(),
                reason=str(cells[9]).strip() if len(cells) > 9 else "",
            ))
        if not header_seen:
            raise ValueError(f"{path}: header row not found")
        if not out:
            raise ValueError(f"{path}: no data rows parsed")
        if unknown_votes:
            warnings.warn(f"{path}: {len(unknown_votes)} unknown vote values, e.g. {unknown_votes[:5]}",
                          stacklevel=2)
        return out
    finally:
        wb.close()
