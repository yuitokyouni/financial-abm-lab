"""大和アセットマネジメント 個別開示パーサ (月次ファイル)。

形式 (2026-07-23 実査): シート名「個別開示　YYYY年M月」、1行目がヘッダ:
  コード / 企業名 / 総会種類(定時総会) / 総会日(YYYYMMDD) / 議案番号('1'|'2.10' 文字列) /
  議案種類(会社提案|株主提案) / 議案分類 / 役員情報 / 判断 / 利益相反 / 賛否判断理由 / 不統一行使
"""

from __future__ import annotations

import unicodedata
import warnings

import openpyxl

from yh010g.schema import (
    UnifiedRecord, map_proposer, map_vote, normalize_date, normalize_sec_code, parse_proposal_no,
)

MANAGER = "daiwa"


def parse_daiwa(path: str) -> list[UnifiedRecord]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        names = [n for n in wb.sheetnames if n.startswith("個別開示")]
        if not names:
            raise ValueError(f"{path}: no 個別開示 sheet (got {wb.sheetnames})")
        out: list[UnifiedRecord] = []
        unknown_votes: list[str] = []
        skipped_special: list[str] = []
        for sn in names:
            ws = wb[sn]
            header_seen = False
            for row in ws.iter_rows(values_only=True):
                cells = [c if c is not None else "" for c in row]
                first = unicodedata.normalize("NFKC", str(cells[0])).strip()
                if not header_seen:
                    if first == "コード":
                        header_seen = True
                    continue
                if first == "":
                    continue
                try:
                    pno, sub = parse_proposal_no(cells[4])
                except ValueError:
                    # 'C.1' 等の非数値議案番号 (種類株主総会系、ごく少数)。
                    # 他社と結合不能なためスキップし記録する
                    skipped_special.append(f"{cells[0]}:{cells[4]}")
                    continue
                vote, known = map_vote(cells[8])
                if not known:
                    unknown_votes.append(str(cells[8]))
                out.append(UnifiedRecord(
                    manager=MANAGER,
                    sec_code=normalize_sec_code(cells[0]),
                    company_name=str(cells[1]).strip(),
                    meeting_date=normalize_date(cells[3]),
                    meeting_type=str(cells[2]).strip(),
                    proposal_no=pno,
                    sub_no=sub,
                    proposer=map_proposer(cells[5]),
                    category=str(cells[6]).strip(),
                    vote=vote,
                    vote_raw=str(cells[8]).strip(),
                    reason=str(cells[10]).strip() if len(cells) > 10 else "",
                ))
            if not header_seen:
                raise ValueError(f"{path}: header row not found in {sn!r}")
        if not out:
            raise ValueError(f"{path}: no data rows parsed")
        if unknown_votes:
            warnings.warn(f"{path}: {len(unknown_votes)} unknown vote values, e.g. {unknown_votes[:5]}",
                          stacklevel=2)
        if skipped_special:
            warnings.warn(f"{path}: skipped {len(skipped_special)} non-numeric proposal rows "
                          f"{skipped_special[:3]}", stacklevel=2)
        return out
    finally:
        wb.close()
