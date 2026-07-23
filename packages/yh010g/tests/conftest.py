"""パーサテスト用フィクスチャ — 実ファイルの構造を再現した合成 xlsx をその場で生成。

実データの抜粋は各社サイト規約 (無断複製禁止) に配慮しコミットしない。
構造の根拠は docs/2026-07-23-YH010g-disclosure-inventory.md の実査記録。
実ファイルでの煙테스트は data/raw/yh010g/ が存在する環境でのみ実行される (test_real_files)。
"""

import openpyxl
import pytest


@pytest.fixture
def mufg_trust_file(tmp_path):
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "議案別議決権行使状況"  # 集計シート (パーサは読まない)
    ws = wb.create_sheet("議決権行使結果")
    ws.append(["会社別議案別行使結果　（2025年4月～2025年6月）"])
    ws.append(["", "", "", "", "", "", "", "", "", "*：顧客ガイドラインによる不統一行使が発生"])
    ws.append(["銘柄コード", "銘柄名称", "総会日", "総会種類", "議案番号", "子議案番号",
               "提案者", "議案分類", "賛否", "理由　（詳細はシート「反対理由詳細」をご参照）"])
    ws.append(["1301", "極洋", "20250625", "定時総会", "1", "0", "会社", "剰余金の処分", "賛成", "特段問題なく、賛成。"])
    ws.append(["1301", "極洋", "20250625", "定時総会", "2", "1", "会社", "取締役の選解任", "反対", "政策保有"])
    ws.append(["1301", "極洋", "20250625", "定時総会", "2", "2", "会社", "取締役の選解任", "賛成*", "特段問題なく、賛成。"])
    ws.append(["9999", "テスト工業", "20250620", "臨時総会", "1", "", "株主", "定款に関する議案", "反対", "株主価値"])
    wb.create_sheet("→反対理由詳細")
    p = tmp_path / "mufg.xlsx"
    wb.save(p)
    return str(p)


@pytest.fixture
def amova_file(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "個別開示"
    ws.append(["議決権行使結果の個別開示（2025年4月-6月）"])
    ws.append(["● 弊社が議決権行使を行った全議案を開示。"])
    ws.append(["● 弊社の議決権等行使指図ガイドラインに基づき行使。"])
    ws.append(["企業コード", "企業名", "総会種類", "総会日", "親議案番号", "子議案番号",
               "議案分類", "提案者", "弊社賛否", "賛否理由", "備考"])
    ws.append(["1301", "極洋", "定時総会", "20250625", "1", "", "剰余金の処分", "会社提案", "賛成", "", ""])
    ws.append(["1301", "極洋", "定時総会", "20250625", "2", "1", "取締役の選解任", "会社提案", "反対", "基準未達", ""])
    ws.append(["9999", "テスト工業", "臨時総会", "2025-06-20", "1", "", "定款一部変更", "株主提案", "賛成", "", ""])
    wb.create_sheet("集計結果")
    p = tmp_path / "amova.xlsx"
    wb.save(p)
    return str(p)


@pytest.fixture
def nissay_file(tmp_path):
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "概要"
    ws = wb.create_sheet("議案別行使結果（6月）")
    ws.append(["ニッセイアセットマネジメント議決権行使結果"])
    ws.append(["報告対象期間：2025年06月01日から2025年06月30日"])
    ws.append([])
    ws.append([])
    ws.append(["（株）極洋(1301)"])
    ws.append(["株主総会開催日　2025年06月25日"])
    ws.append(["議案\n番号", "候補者\n番号", "議案区分", "提案", "判断", "主な判断理由"])
    ws.append(["1", "", "利益処分に関する議案", "会社", "賛成", "判断基準に基づき賛成"])
    ws.append(["2", "1", "取締役の選任等に関する議案", "会社", "反対", "独立性に関する基準に該当"])
    ws.append([])
    ws.append([])
    ws.append(["テスト工業（株）(9999)"])
    ws.append(["株主総会開催日　2025年06月20日"])
    ws.append(["議案\n番号", "候補者\n番号", "議案区分", "提案", "判断", "主な判断理由"])
    ws.append(["1", "", "定款変更に関する議案", "株主", "賛成", "株主価値向上に資する"])
    wb.create_sheet("議案別行使結果_REIT")
    p = tmp_path / "nissay.xlsx"
    wb.save(p)
    return str(p)
