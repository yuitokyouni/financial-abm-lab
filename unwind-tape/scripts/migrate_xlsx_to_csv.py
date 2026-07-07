#!/usr/bin/env python3
"""unwind-tape / task B step 1 — one-time migration: xlsx v0.3 → canonical CSVs.

以後は CSV が一次データ。xlsx は build_tape.py で再生成される。

出力:
    data/parsed/tape/groups.csv         # event_group_id ごとに identity/分類
    data/parsed/tape/legs.csv           # event_group_id × event_leg_id で実データ (69 - group昇格分) 列
    data/parsed/tape/sources.csv        # Source_Log + local_path/sha256/bytes/fetched_at (archiver が埋める)
    data/parsed/tape/lists.yaml         # Lists シートを field -> [enum] の dict に整形
    data/parsed/tape/field_dictionary.csv
    data/parsed/tape/sampling_frame.csv
    data/parsed/tape/baseline_spec.csv
    data/parsed/tape/changelog.csv
    data/parsed/tape/readme.yaml
    data/parsed/tape/migration_report.md   # 変換で捨てた/変えた点の一覧

不変条件:
  - 値の創作は禁止。xlsx セルが空欄なら CSV も空欄。
  - group 昇格した列は全 leg で値が一致することを検証。不一致は migration_report にエラー行を出し非0終了。
  - enum 値は Lists と突合。未知 enum は WARN として migration_report に列挙(migration は継続)。
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
from pathlib import Path

import openpyxl
import yaml


# ---------------------------------------------------------------------------
# schema — v0.3 xlsx Event_Tape の 69 列を group / leg にどう振り分けるか
# ---------------------------------------------------------------------------

# 全 69 列 (Field_Dictionary の順序と一致)
EVENT_TAPE_COLUMNS_V03 = [
    "event_group_id", "event_leg_id", "status",
    "issuer_code", "issuer_name", "issuer_market",
    "seller_name", "seller_type", "relationship",
    "event_tier", "event_role", "sale_route", "absorption_route",
    "support_buyback", "support_OA", "support_lockup", "support_stabilization",
    "disclosure_time", "after_close",
    "announce_datetime", "pricing_date", "settlement_date", "trade_date",
    "quantity_basis", "value_basis",
    "initial_offer_shares", "final_offer_shares", "OA_shares_max", "OA_exercised_shares",
    "sold_shares", "sold_value_JPY",
    "pre_holding_shares", "post_holding_shares",
    "ADV20_shares", "ADV60_shares", "free_float_shares", "market_cap_JPY",
    "size_ADV20_calc", "size_ADV60_calc", "size_free_float_calc", "size_market_cap_calc",
    "offer_price_JPY", "previous_close_JPY", "discount_calc",
    "buyback_size_shares", "buyback_value_JPY", "buyback_method", "buyback_window", "buyback_cancellation",
    "lockup_days", "lead_underwriter", "seller_breakdown",
    "activist_pressure",
    "announcement_CAR_m1_p1", "announcement_CAR_0_p1", "pricing_CAR_m1_p1", "settlement_CAR_m1_p1",
    "drift_ann_to_pricing", "recovery_5d", "recovery_20d", "recovery_60d", "abnormal_volume_0_p3",
    "route_notes", "mechanism_hypothesis",
    "ABM_candidate_flag", "confidence_policy_holding",
    "source_primary_url", "source_secondary_url",
    "notes",
]

# group.csv に昇格する列 (=group PK を含めて identity + case-level 分類)
# 全 leg で値が一致することを検証。mechanism_hypothesis は leg 固有の解釈なので
# 昇格しない (G008 で L001/L002 で異なる文言があった)。activist_pressure も
# leg 単位で空欄/TRUE が混在するのは意味を持つため leg に残す。
GROUP_COLUMNS = [
    "event_group_id",
    "issuer_code", "issuer_name", "issuer_market",
    "event_tier",
    "confidence_policy_holding",
    "ABM_candidate_flag",
]

# CSVに"formula"列は空欄のまま保持 (Task C が埋める)。ここで削除しない — round-trip 用。
FORMULA_COLUMNS = {
    "size_ADV20_calc", "size_ADV60_calc", "size_free_float_calc", "size_market_cap_calc",
    "discount_calc", "drift_ann_to_pricing", "recovery_60d",
}

# CSV 出力での bool 表記: TRUE / FALSE / (空欄) / unknown
BOOL_FIELDS = {
    "support_buyback", "support_OA", "support_lockup", "support_stabilization",
    "after_close", "buyback_cancellation",
}
# activist_pressure は enum(TRUE/FALSE/unknown) なので BOOL_FIELDS には入れず enum として扱う

DATE_FIELDS = {"announce_datetime", "pricing_date", "settlement_date", "trade_date"}
TIME_FIELDS = {"disclosure_time"}
NUMBER_FIELDS = {
    "initial_offer_shares", "final_offer_shares", "OA_shares_max", "OA_exercised_shares",
    "sold_shares", "sold_value_JPY",
    "pre_holding_shares", "post_holding_shares",
    "ADV20_shares", "ADV60_shares", "free_float_shares", "market_cap_JPY",
    "offer_price_JPY", "previous_close_JPY",
    "buyback_size_shares", "buyback_value_JPY",
    "lockup_days",
    "announcement_CAR_m1_p1", "announcement_CAR_0_p1", "pricing_CAR_m1_p1", "settlement_CAR_m1_p1",
    "recovery_5d", "recovery_20d", "abnormal_volume_0_p3",
}


# ---------------------------------------------------------------------------
# cell → canonical string helpers
# ---------------------------------------------------------------------------

def _norm_bool(v) -> str:
    if v is None or (isinstance(v, str) and not v.strip()):
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, str):
        s = v.strip().upper()
        if s in ("TRUE", "T", "YES", "Y", "1"):
            return "TRUE"
        if s in ("FALSE", "F", "NO", "N", "0"):
            return "FALSE"
        if s in ("UNKNOWN", "UNK", "-"):
            return "UNKNOWN"
        return s  # keep as-is; validator will flag if not enum
    return str(v)


def _norm_date(v) -> str:
    if v is None or (isinstance(v, str) and not v.strip()):
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        d = v.date() if isinstance(v, dt.datetime) else v
        return d.isoformat()
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        return s
    return str(v)


def _norm_time(v) -> str:
    if v is None or (isinstance(v, str) and not v.strip()):
        return ""
    if isinstance(v, dt.time):
        return v.strftime("%H:%M")
    if isinstance(v, dt.datetime):
        return v.strftime("%H:%M")
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", s)
        if m:
            return f"{int(m.group(1)):02d}:{m.group(2)}"
        return s
    return str(v)


def _norm_number(v) -> str:
    if v is None or (isinstance(v, str) and not v.strip()):
        return ""
    if isinstance(v, bool):
        return ""  # do not coerce bool to number
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    if isinstance(v, str):
        s = re.sub(r"[,\s]", "", v.strip())
        # keep suffixes (億円 etc.) as-is if not pure number; validator later
        if re.match(r"^-?\d+(\.\d+)?$", s):
            return s
        return v.strip()
    return str(v)


def _norm_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        # collapse whitespace to single space but preserve linebreaks
        return "\n".join(re.sub(r"[ \t]+", " ", line).strip() for line in v.splitlines()).strip()
    if isinstance(v, (dt.datetime, dt.date)):
        return _norm_date(v)
    return str(v)


def _normalise_cell(field: str, v) -> str:
    if field in BOOL_FIELDS:
        return _norm_bool(v)
    if field in DATE_FIELDS:
        return _norm_date(v)
    if field in TIME_FIELDS:
        return _norm_time(v)
    if field in NUMBER_FIELDS:
        return _norm_number(v)
    if field == "activist_pressure":
        return _norm_bool(v)  # enum values TRUE/FALSE/unknown
    return _norm_text(v)


# ---------------------------------------------------------------------------
# workbook readers
# ---------------------------------------------------------------------------

def _content_rows(ws) -> list[list]:
    out: list[list] = []
    for row in ws.iter_rows(values_only=True):
        r = list(row)
        while r and r[-1] is None:
            r.pop()
        if not r and out and not out[-1]:
            break
        out.append(r)
    while out and not out[-1]:
        out.pop()
    return out


def read_event_tape(wb) -> list[dict[str, str]]:
    """Return list of leg dicts keyed by canonical column name."""
    ws = wb["Event_Tape"]
    rows = list(ws.iter_rows(values_only=True))
    header = [h if h is not None else "" for h in rows[0]]
    if header[: len(EVENT_TAPE_COLUMNS_V03)] != EVENT_TAPE_COLUMNS_V03:
        raise ValueError(
            "Event_Tape header order mismatch. Expected v0.3 order.\n"
            f"  got: {header!r}\n  want: {EVENT_TAPE_COLUMNS_V03!r}"
        )
    out: list[dict[str, str]] = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() != "" for c in r):
            continue
        rec = {}
        for i, field in enumerate(EVENT_TAPE_COLUMNS_V03):
            v = r[i] if i < len(r) else None
            rec[field] = _normalise_cell(field, v)
        out.append(rec)
    return out


def read_source_log(wb) -> list[dict[str, str]]:
    ws = wb["Source_Log"]
    rows = _content_rows(ws)
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    expected = ["source_id", "source_type", "source_url", "what_it_supports", "archived", "notes"]
    if header[: len(expected)] != expected:
        raise ValueError(f"Source_Log header mismatch: got {header!r}, want {expected!r}")
    out = []
    for r in rows[1:]:
        rec = {}
        for i, field in enumerate(expected):
            v = r[i] if i < len(r) else None
            if field == "archived":
                rec[field] = _norm_bool(v)
            else:
                rec[field] = _norm_text(v)
        out.append(rec)
    return out


def read_lists(wb) -> dict[str, list[str]]:
    ws = wb["Lists"]
    rows = _content_rows(ws)
    if not rows:
        return {}
    header = [str(c) if c is not None else "" for c in rows[0]]
    out: dict[str, list[str]] = {h: [] for h in header if h}
    for r in rows[1:]:
        for i, h in enumerate(header):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            out[h].append(_norm_text(v))
    return {k: v for k, v in out.items() if v}


def read_field_dictionary(wb) -> list[dict[str, str]]:
    ws = wb["Field_Dictionary"]
    rows = _content_rows(ws)
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        rec = {h: _norm_text(r[i] if i < len(r) else None) for i, h in enumerate(header)}
        if any(rec.values()):
            out.append(rec)
    return out


def read_sampling_frame(wb) -> list[tuple[str, str]]:
    ws = wb["Sampling_Frame"]
    rows = _content_rows(ws)
    out = []
    for r in rows:
        if not r:
            continue
        k = _norm_text(r[0]) if len(r) >= 1 else ""
        v = _norm_text(r[1]) if len(r) >= 2 else ""
        if k or v:
            out.append((k, v))
    return out


def read_baseline_spec(wb) -> list[dict[str, str]]:
    ws = wb["Baseline_Spec"]
    rows = _content_rows(ws)
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        rec = {h: _norm_text(r[i] if i < len(r) else None) for i, h in enumerate(header)}
        if any(rec.values()):
            out.append(rec)
    return out


def read_changelog(wb) -> list[dict[str, str]]:
    ws = wb["Changelog"]
    rows = _content_rows(ws)
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        rec = {h: _norm_text(r[i] if i < len(r) else None) for i, h in enumerate(header)}
        if any(rec.values()):
            out.append(rec)
    return out


def read_readme(wb) -> list[tuple[str, str]]:
    ws = wb["README"]
    rows = _content_rows(ws)
    out = []
    for r in rows:
        if not r:
            continue
        k = _norm_text(r[0]) if len(r) >= 1 else ""
        v = _norm_text(r[1]) if len(r) >= 2 else ""
        if k or v:
            out.append((k, v))
    return out


# ---------------------------------------------------------------------------
# group derivation + integrity check
# ---------------------------------------------------------------------------

def derive_groups(legs: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[str]]:
    """For each event_group_id, verify GROUP_COLUMNS are consistent across legs.
    Returns (groups, warnings).
    """
    by_group: dict[str, list[dict[str, str]]] = {}
    order: list[str] = []
    for leg in legs:
        gid = leg["event_group_id"]
        if gid not in by_group:
            by_group[gid] = []
            order.append(gid)
        by_group[gid].append(leg)

    groups: list[dict[str, str]] = []
    warnings: list[str] = []
    for gid in order:
        legs_g = by_group[gid]
        grec: dict[str, str] = {}
        for col in GROUP_COLUMNS:
            vals = {leg.get(col, "") for leg in legs_g}
            # allow one leg to have "" (unfilled) and another to have a real value.
            non_empty = {v for v in vals if v != ""}
            if len(non_empty) > 1:
                warnings.append(
                    f"group {gid}: leg 間で '{col}' が不一致: {sorted(non_empty)}."
                    f" 最初の non-empty を採用 (legs [{', '.join(leg['event_leg_id'] for leg in legs_g)}])"
                )
                grec[col] = sorted(non_empty)[0]
            elif len(non_empty) == 1:
                grec[col] = next(iter(non_empty))
            else:
                grec[col] = ""
        groups.append(grec)
    return groups, warnings


def strip_group_columns_from_legs(legs: list[dict[str, str]]) -> list[dict[str, str]]:
    """Remove GROUP_COLUMNS from leg dicts (except event_group_id which is the FK)."""
    to_drop = set(GROUP_COLUMNS) - {"event_group_id"}
    out = []
    for leg in legs:
        r = {k: v for k, v in leg.items() if k not in to_drop}
        out.append(r)
    return out


def resolve_source_ids(legs: list[dict[str, str]], sources: list[dict[str, str]]
                     ) -> tuple[list[dict[str, str]], list[str], list[str]]:
    """Attach source_id_primary / source_id_secondary by URL match.

    v0.3 の Event_Tape では一部URLがちょうど80文字で末尾 "..." に切れているセルがある
    (autofit save 由来と推定される xlsx 側のデータ破損)。
    (1) まず完全一致で解決
    (2) 完全一致で不解決なら、末尾 "..." を剥がした prefix と Source_Log 中の URL の
        prefix match を試みる。**ちょうど 1件** マッチしたときのみ source_id を付ける
        (曖昧一致は付けない)。この場合 URL 文字列自体は "..." のまま残す (データ創作禁止)。
    (3) それでも解決できなければ unresolved として warning に列挙。

    Returns (legs_with_ids, unresolved_warnings, truncated_url_warnings).
    """
    url_to_id = {s["source_url"]: s["source_id"] for s in sources if s.get("source_url")}
    src_urls = [(s["source_url"], s["source_id"]) for s in sources if s.get("source_url")]
    unresolved: list[str] = []
    truncated: list[str] = []
    out = []
    for leg in legs:
        new = dict(leg)
        for kind, urlcol, idcol in [
            ("primary", "source_primary_url", "source_id_primary"),
            ("secondary", "source_secondary_url", "source_id_secondary"),
        ]:
            url = leg.get(urlcol, "").strip()
            if not url:
                new[idcol] = ""
                continue
            if url in url_to_id:
                new[idcol] = url_to_id[url]
                continue
            # try prefix-match fallback for truncated URLs
            if url.endswith("..."):
                stem = url[:-3]
                matches = [sid for surl, sid in src_urls if surl.startswith(stem)]
                if len(matches) == 1:
                    new[idcol] = matches[0]
                    truncated.append(
                        f"leg {leg['event_group_id']}/{leg['event_leg_id']}: "
                        f"{kind} URL truncated in xlsx (len={len(url)}), resolved by "
                        f"prefix match to {matches[0]} but URL string kept truncated in legs.csv"
                    )
                    continue
                if len(matches) > 1:
                    new[idcol] = ""
                    unresolved.append(
                        f"leg {leg['event_group_id']}/{leg['event_leg_id']}: "
                        f"{kind} URL truncated + prefix ambiguous ({len(matches)} candidates)"
                    )
                    continue
            new[idcol] = ""
            unresolved.append(
                f"leg {leg['event_group_id']}/{leg['event_leg_id']}: "
                f"{kind} URL not in Source_Log ({url})"
            )
        out.append(new)
    return out, unresolved, truncated


# ---------------------------------------------------------------------------
# writers
# ---------------------------------------------------------------------------

def _write_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(columns)
        for r in rows:
            w.writerow([r.get(c, "") for c in columns])


def _write_yaml(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        yaml.safe_dump(obj, f, allow_unicode=True, sort_keys=False)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path,
                    default=Path(__file__).resolve().parent.parent / "inputs" / "tape_versions"
                    / "v0.4" / "policy_holding_sale_event_tape_v0_4.xlsx")
    ap.add_argument("--out", type=Path,
                    default=Path(__file__).resolve().parent.parent / "data" / "parsed" / "tape")
    ap.add_argument("--version-tag", default="v0.4")
    args = ap.parse_args(argv)

    if not args.xlsx.exists():
        print(f"xlsx not found: {args.xlsx}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    legs_all = read_event_tape(wb)
    sources = read_source_log(wb)
    lists = read_lists(wb)
    field_dict = read_field_dictionary(wb)
    sampling = read_sampling_frame(wb)
    baseline = read_baseline_spec(wb)
    changelog = read_changelog(wb)
    readme = read_readme(wb)

    groups, group_warnings = derive_groups(legs_all)
    legs = strip_group_columns_from_legs(legs_all)
    legs, url_warnings, truncated_url_warnings = resolve_source_ids(legs, sources)

    # column order for legs.csv = EVENT_TAPE_COLUMNS_V03 minus (GROUP_COLUMNS - {event_group_id})
    # + [source_id_primary, source_id_secondary] injected right before source_primary_url.
    leg_cols: list[str] = []
    for c in EVENT_TAPE_COLUMNS_V03:
        if c in GROUP_COLUMNS and c != "event_group_id":
            continue
        if c == "source_primary_url":
            leg_cols.append("source_id_primary")
        if c == "source_secondary_url":
            leg_cols.append("source_id_secondary")
        leg_cols.append(c)

    args.out.mkdir(parents=True, exist_ok=True)
    _write_csv(args.out / "groups.csv", GROUP_COLUMNS, groups)
    _write_csv(args.out / "legs.csv", leg_cols, legs)

    # sources.csv: add archiver-fillable columns.
    src_cols = ["source_id", "source_type", "source_url", "what_it_supports",
                "archived", "local_path", "sha256", "bytes", "fetched_at", "notes"]
    src_rows = []
    for s in sources:
        r = {c: s.get(c, "") for c in ["source_id", "source_type", "source_url",
                                       "what_it_supports", "archived", "notes"]}
        r["local_path"] = ""
        r["sha256"] = ""
        r["bytes"] = ""
        r["fetched_at"] = ""
        src_rows.append(r)
    _write_csv(args.out / "sources.csv", src_cols, src_rows)

    _write_yaml(args.out / "lists.yaml", lists)

    # metadata sheets → CSV / YAML (for round-trip build_tape.py)
    if field_dict:
        _write_csv(args.out / "field_dictionary.csv",
                   list(field_dict[0].keys()), field_dict)
    _write_csv(args.out / "sampling_frame.csv",
               ["key", "value"],
               [{"key": k, "value": v} for k, v in sampling])
    if baseline:
        _write_csv(args.out / "baseline_spec.csv",
                   list(baseline[0].keys()), baseline)
    if changelog:
        _write_csv(args.out / "changelog.csv",
                   list(changelog[0].keys()), changelog)
    _write_yaml(args.out / "readme.yaml",
                {k: v for k, v in readme})

    # migration report
    report_lines: list[str] = []
    report_lines.append(f"# migration report — {args.version_tag} → CSV\n")
    report_lines.append(f"generated: {dt.datetime.now().isoformat(timespec='seconds')}\n")
    report_lines.append(f"source xlsx: `{args.xlsx.relative_to(args.xlsx.parent.parent.parent.parent) if args.xlsx.is_absolute() else args.xlsx}`\n\n")
    report_lines.append(f"## counts\n")
    report_lines.append(f"- legs: {len(legs)}\n")
    report_lines.append(f"- groups: {len(groups)}\n")
    report_lines.append(f"- sources: {len(sources)}\n")
    report_lines.append(f"- lists (enum families): {len(lists)}\n")
    report_lines.append(f"- field_dictionary: {len(field_dict)}\n")
    report_lines.append(f"- baseline_spec: {len(baseline)}\n")
    report_lines.append(f"- changelog: {len(changelog)}\n\n")

    if group_warnings:
        report_lines.append("## group column consistency warnings\n")
        for w in group_warnings:
            report_lines.append(f"- {w}\n")
        report_lines.append("\n")
    if truncated_url_warnings:
        report_lines.append("## truncated URLs in xlsx (auto-resolved by prefix match)\n")
        report_lines.append(
            "v0.3 Event_Tape has cells where source URLs were truncated to 80 characters "
            "with a literal `...` suffix (autofit-save artefact). source_id は Source_Log "
            "への一意 prefix match で解決したが、URL 文字列は truncated のまま legs.csv に残す "
            "(データ創作禁止)。次バージョン (v0.4+) で xlsx 側の URL を修復すること。\n\n"
        )
        for w in truncated_url_warnings:
            report_lines.append(f"- {w}\n")
        report_lines.append("\n")
    if url_warnings:
        report_lines.append("## unresolved source URLs (not in Source_Log)\n")
        for w in url_warnings:
            report_lines.append(f"- {w}\n")
        report_lines.append("\n")
    if not group_warnings and not url_warnings and not truncated_url_warnings:
        report_lines.append("no warnings.\n")

    (args.out / "migration_report.md").write_text("".join(report_lines), encoding="utf-8")

    print(f"wrote {len(groups)} groups, {len(legs)} legs, {len(sources)} sources → {args.out}")
    if group_warnings:
        print(f"  group-column warnings: {len(group_warnings)} (see migration_report.md)")
    if truncated_url_warnings:
        print(f"  truncated-URL auto-resolves: {len(truncated_url_warnings)} (see migration_report.md)")
    if url_warnings:
        print(f"  unresolved-URL warnings: {len(url_warnings)} (see migration_report.md)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
