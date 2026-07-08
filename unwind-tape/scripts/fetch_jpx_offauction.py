#!/usr/bin/env python3
"""unwind-tape / task A — JPX 立会外取引情報 daily fetch.

3ページを毎日キャプチャする:
  1) ToSTNeT 超大口約定情報 (ToSTNeT-1, xlsx per publication_date)
  2) 立会外分売情報 (HTML rowspan=2 表、過去2週間分)
  3) 自己株式立会外買付取引情報 (HTML 単純表、過去2週間分)

不変条件:
  - raw をそのまま data/raw/jpx_offauction/{page}/{YYYY-MM-DD}/ に保存
  - manifest.jsonl に sha256 込みで追記
  - parsed CSV は natural_key で idempotent に append (掲載列は全て保持)
  - 想定カラムと不一致なら raw のみ保存し schema_mismatch として非0終了
  - 欠損の創作は絶対しない。取得失敗/schema不整合は gaps_report.md に列挙。

依存: requests, openpyxl, PyYAML, stdlib のみ。他パッケージへ import しない。
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import logging
import re
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin
from zoneinfo import ZoneInfo

import openpyxl
import requests
import yaml


# ---------------------------------------------------------------------------
# util
# ---------------------------------------------------------------------------

def _sha256(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _load_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def _now(tz: ZoneInfo) -> datetime:
    # zoneinfo-aware now (allowed; Date.now() sandbox note is workflow-only).
    return datetime.now(tz)


def _iso_date_from_yyyymmdd(s: str) -> str:
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"


def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def _strip_yen_suffix_int(cell: str) -> str:
    # "600,000株" -> "600000", "1,424円" -> "1424"
    if cell is None:
        return ""
    v = re.sub(r"[,\s]", "", str(cell))
    v = re.sub(r"(円|株)$", "", v)
    return v


ISSUE_CODE_RE = re.compile(r"（([0-9A-Za-z]{4,5})）\s*$")
ISSUE_NAME_SPLIT_RE = re.compile(r"^(?P<name>.+?)（(?P<code>[0-9A-Za-z]{4,5})）\s*$")


def _split_name_code(cell: str) -> tuple[str, str]:
    """'（株）ヤマウラ　株式（1780）' -> ('（株）ヤマウラ　株式', '1780').

    抽出できないときは (原文, '') を返す。gaps_report に載る。
    """
    s = _norm_ws(cell)
    m = ISSUE_NAME_SPLIT_RE.match(s)
    if not m:
        return s, ""
    return m.group("name").strip(), m.group("code").strip()


# ---------------------------------------------------------------------------
# HTTP with retry
# ---------------------------------------------------------------------------

class Http:
    def __init__(self, ua: str, timeout: float, retries: int, base: float, factor: float):
        self.ua = ua
        self.timeout = timeout
        self.retries = retries
        self.base = base
        self.factor = factor
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": ua, "Accept-Language": "ja,en;q=0.5"})

    def get(self, url: str, *, log: logging.Logger) -> bytes:
        last_exc: Exception | None = None
        for attempt in range(self.retries + 1):
            try:
                r = self.session.get(url, timeout=self.timeout)
                r.raise_for_status()
                return r.content
            except Exception as e:
                last_exc = e
                if attempt >= self.retries:
                    break
                wait = self.base * (self.factor ** attempt)
                log.warning("GET %s failed (%s); retry %d/%d in %.1fs", url, e, attempt + 1, self.retries, wait)
                time.sleep(wait)
        assert last_exc is not None
        raise last_exc


# ---------------------------------------------------------------------------
# manifest
# ---------------------------------------------------------------------------

@dataclass
class ManifestEntry:
    captured_at: str
    page: str
    date_key: str            # YYYY-MM-DD (publication_date or capture_date)
    source_url: str
    local_path: str
    sha256: str
    bytes: int
    rows_parsed: int | None
    status: str              # ok | schema_mismatch | fetch_error | skipped_duplicate
    detail: str | None = None


def _append_manifest(manifest_path: Path, entry: ManifestEntry) -> None:
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    with manifest_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry.__dict__, ensure_ascii=False) + "\n")


# ---------------------------------------------------------------------------
# freshness alarm (2026-07-08 review 指摘)
# ---------------------------------------------------------------------------
# 掲載保持は「過去2週間」しかない。cron/launchd の発火漏れ・構造変化を
# 誰も見ていないと、猶予が尽きて欠測期間が生まれてから気付くことになる。
# manifest.jsonl の最新 status=ok の date_key が営業日換算で N 日以上
# 古ければ ERROR ログを出し、非0終了させる。祝日は考慮しない
# (月〜金の単純カウント。アラーム閾値としては十分な精度)。

def _weekdays_between(d: date, today: date) -> int:
    """d(exclusive) から today(inclusive) までの月〜金日数。祝日非考慮。"""
    count = 0
    cur = d
    while cur < today:
        cur += timedelta(days=1)
        if cur.weekday() < 5:
            count += 1
    return count


def _latest_ok_date_key(manifest_path: Path) -> str | None:
    latest: str | None = None
    if not manifest_path.exists():
        return None
    with manifest_path.open("r", encoding="utf-8") as f:
        for line in f:
            try:
                e = json.loads(line)
            except json.JSONDecodeError:
                continue
            if e.get("status") != "ok":
                continue
            dk = e.get("date_key")
            if dk and (latest is None or dk > latest):
                latest = dk
    return latest


def _check_freshness(root: Path, storage: dict, pages: list[str], now: datetime,
                     max_stale_business_days: int, log: logging.Logger) -> bool:
    """各ページの manifest.jsonl から最新 status=ok の date_key を見て、
    今日との営業日差が閾値を超えていたら ERROR ログを出す。
    Returns True if any page is stale (呼び出し元で exit_code に反映すること)。
    """
    today = now.date()
    any_stale = False
    for page in pages:
        manifest_path = root / storage["raw_root"] / page / storage["manifest_filename"]
        latest = _latest_ok_date_key(manifest_path)
        if latest is None:
            log.error("freshness alarm: page=%s has no status=ok manifest entry at all "
                      "(never successfully captured?)", page)
            any_stale = True
            continue
        try:
            y, m, d = (int(x) for x in latest.split("-"))
            latest_date = date(y, m, d)
        except ValueError:
            log.error("freshness alarm: page=%s latest date_key %r unparseable", page, latest)
            any_stale = True
            continue
        gap = _weekdays_between(latest_date, today)
        if gap > max_stale_business_days:
            log.error(
                "freshness alarm: page=%s latest captured date_key=%s is %d business days "
                "old (threshold=%d). JPX 掲載保持は過去2週間のみ — 発火漏れ・構造変化を確認すること。",
                page, latest, gap, max_stale_business_days,
            )
            any_stale = True
        else:
            log.info("freshness ok: page=%s latest date_key=%s (%d business days old)",
                     page, latest, gap)
    return any_stale


def _manifest_has_sha256(manifest_path: Path, sha256: str, status_ok_only: bool = True) -> bool:
    if not manifest_path.exists():
        return False
    with manifest_path.open("r", encoding="utf-8") as f:
        for line in f:
            try:
                e = json.loads(line)
            except json.JSONDecodeError:
                continue
            if e.get("sha256") == sha256 and (not status_ok_only or e.get("status") == "ok"):
                return True
    return False


# ---------------------------------------------------------------------------
# HTML table extraction
# ---------------------------------------------------------------------------

class _TableCollector(HTMLParser):
    """Collect <table>...</table>s: rows -> list[(tag, [cells])], tag in {'th','td'}.

    Whitespace is normalised. <br> becomes ' '. Nested tags are ignored.
    """

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[tuple[str, list[str]]]] = []
        self._in_table = False
        self._current_table: list[tuple[str, list[str]]] | None = None
        self._current_row_kind: str | None = None      # 'th' | 'td' if seen
        self._current_row_cells: list[str] = []
        self._current_cell_buf: list[str] = []
        self._in_cell = False
        self._cell_tag: str | None = None
        self._in_tr = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "table":
            self._in_table = True
            self._current_table = []
            return
        if not self._in_table:
            return
        if tag == "tr":
            self._in_tr = True
            self._current_row_kind = None
            self._current_row_cells = []
            return
        if tag in ("th", "td") and self._in_tr:
            self._in_cell = True
            self._cell_tag = tag
            self._current_cell_buf = []
            if self._current_row_kind is None:
                self._current_row_kind = tag
            elif self._current_row_kind == "th" and tag == "td":
                self._current_row_kind = "td"
            return
        if tag == "br" and self._in_cell:
            self._current_cell_buf.append(" ")

    def handle_endtag(self, tag: str) -> None:
        if tag == "table":
            if self._in_table and self._current_table is not None:
                self.tables.append(self._current_table)
            self._in_table = False
            self._current_table = None
            self._in_tr = False
            return
        if not self._in_table:
            return
        if tag == "tr":
            if self._in_tr and self._current_table is not None:
                kind = self._current_row_kind or "td"
                self._current_table.append((kind, self._current_row_cells))
            self._in_tr = False
            self._current_row_kind = None
            self._current_row_cells = []
            return
        if tag in ("th", "td") and self._in_cell:
            text = _norm_ws("".join(self._current_cell_buf))
            self._current_row_cells.append(text)
            self._in_cell = False
            self._cell_tag = None
            self._current_cell_buf = []

    def handle_data(self, data: str) -> None:
        if self._in_cell:
            self._current_cell_buf.append(data)


def _extract_tables(html_bytes: bytes) -> list[list[tuple[str, list[str]]]]:
    p = _TableCollector()
    p.feed(html_bytes.decode("utf-8", errors="replace"))
    return p.tables


def _flatten_header_cells(rows: list[tuple[str, list[str]]]) -> list[str]:
    """Concatenate all header-row (th) cell texts in encounter order."""
    hs: list[str] = []
    for kind, cells in rows:
        if kind != "th":
            continue
        hs.extend(cells)
    return hs


# ---------------------------------------------------------------------------
# ToSTNeT xlsx handler
# ---------------------------------------------------------------------------

@dataclass
class PageResult:
    page: str
    ok_files: int = 0
    schema_mismatches: list[str] = field(default_factory=list)
    fetch_errors: list[str] = field(default_factory=list)
    new_rows: int = 0
    duplicates_skipped: int = 0


def _fetch_page_tostnet(cfg_page: dict[str, Any], http: Http, storage: dict[str, str],
                       now_tz: datetime, root: Path, log: logging.Logger) -> PageResult:
    page = "tostnet_large_lots"
    res = PageResult(page=page)
    raw_dir_root = root / storage["raw_root"] / page
    parsed_csv = root / storage["parsed_root"] / f"{page}.csv"
    manifest = raw_dir_root / storage["manifest_filename"]
    raw_dir_root.mkdir(parents=True, exist_ok=True)
    parsed_csv.parent.mkdir(parents=True, exist_ok=True)

    # 1. fetch index and enumerate xlsx links
    try:
        html_bytes = http.get(cfg_page["index_url"], log=log)
    except Exception as e:
        log.error("tostnet index fetch failed: %s", e)
        res.fetch_errors.append(f"index {cfg_page['index_url']}: {e}")
        return res

    href_re = re.compile(cfg_page["file_href_regex"])
    links: list[tuple[str, str]] = []  # (abs_url, YYYYMMDD)
    for m in href_re.finditer(html_bytes.decode("utf-8", errors="replace")):
        href = m.group(1)
        yyyymmdd = m.group(2)
        abs_url = urljoin(cfg_page["index_url"], href)
        links.append((abs_url, yyyymmdd))
    links = sorted(set(links), key=lambda x: x[1])
    log.info("tostnet index: %d xlsx link(s)", len(links))

    if not links:
        # not necessarily an error — the page may be empty on non-business days.
        # But if we expected data (weekday), record a fetch_error so it surfaces.
        res.fetch_errors.append("tostnet index has no xlsx links (page structure change?)")

    expected_cols = cfg_page["expected_columns"]

    for abs_url, yyyymmdd in links:
        pub_iso = _iso_date_from_yyyymmdd(yyyymmdd)
        raw_dir = raw_dir_root / pub_iso
        raw_dir.mkdir(parents=True, exist_ok=True)
        local_path = raw_dir / abs_url.rsplit("/", 1)[-1]

        # download
        try:
            xlsx_bytes = http.get(abs_url, log=log)
        except Exception as e:
            log.error("tostnet xlsx fetch failed %s: %s", abs_url, e)
            res.fetch_errors.append(f"{abs_url}: {e}")
            _append_manifest(manifest, ManifestEntry(
                captured_at=now_tz.isoformat(timespec="seconds"),
                page=page, date_key=pub_iso, source_url=abs_url,
                local_path=str(local_path.relative_to(root)), sha256="", bytes=0,
                rows_parsed=None, status="fetch_error", detail=str(e),
            ))
            continue

        sha = _sha256(xlsx_bytes)
        # idempotency: same sha256 already recorded as ok?
        if local_path.exists() and _manifest_has_sha256(manifest, sha):
            log.info("tostnet %s already captured (sha256 match) — skip", pub_iso)
            _append_manifest(manifest, ManifestEntry(
                captured_at=now_tz.isoformat(timespec="seconds"),
                page=page, date_key=pub_iso, source_url=abs_url,
                local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(xlsx_bytes),
                rows_parsed=None, status="skipped_duplicate", detail=None,
            ))
            res.duplicates_skipped += 1
            continue

        # write raw first (always)
        local_path.write_bytes(xlsx_bytes)

        # parse
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
            sheet_name = cfg_page.get("sheet_name") or wb.sheetnames[0]
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                raise ValueError("empty sheet")
            header_list = [str(c) if c is not None else "" for c in header]
            if header_list != expected_cols:
                raise ValueError(f"header mismatch: got {header_list!r}, expected {expected_cols!r}")
            data_rows: list[list[Any]] = []
            for row in rows_iter:
                if row is None:
                    continue
                if all(v is None or v == "" for v in row):
                    continue
                data_rows.append(list(row))
        except Exception as e:
            log.error("tostnet parse failed %s: %s", pub_iso, e)
            res.schema_mismatches.append(f"{pub_iso}: {e}")
            _append_manifest(manifest, ManifestEntry(
                captured_at=now_tz.isoformat(timespec="seconds"),
                page=page, date_key=pub_iso, source_url=abs_url,
                local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(xlsx_bytes),
                rows_parsed=None, status="schema_mismatch", detail=str(e),
            ))
            continue

        # merge into master CSV, keyed on publication_date (whole file dedupe).
        rewritten = _merge_tostnet_rows(parsed_csv, header_list, data_rows, pub_iso)
        res.new_rows += rewritten
        res.ok_files += 1

        _append_manifest(manifest, ManifestEntry(
            captured_at=now_tz.isoformat(timespec="seconds"),
            page=page, date_key=pub_iso, source_url=abs_url,
            local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(xlsx_bytes),
            rows_parsed=len(data_rows), status="ok", detail=None,
        ))
    return res


def _merge_tostnet_rows(csv_path: Path, header: list[str], rows: list[list[Any]], pub_iso: str) -> int:
    """Whole-publication_date replace: rows with matching publication_date are removed
    from existing CSV, then new rows are appended. Returns count of new rows added.

    A row-hash column is added at end for downstream inspection.
    """
    # Include an ISO-formatted publication_date column derived from the raw 公表日 field
    # so downstream tooling doesn't have to re-parse the '20260706' int form.
    out_header = list(header) + ["_publication_date_iso", "_row_hash"]

    existing_rows: list[list[str]] = []
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8", newline="") as f:
            r = csv.reader(f)
            existing_header = next(r, None)
            if existing_header != out_header:
                # header change — write header only if empty; else refuse.
                if existing_header is None:
                    pass
                else:
                    raise ValueError(
                        f"CSV header mismatch on {csv_path}: file has {existing_header!r} "
                        f"vs expected {out_header!r}"
                    )
            for row in r:
                existing_rows.append(row)

    # drop existing rows for this publication_date
    pub_idx = out_header.index("_publication_date_iso")
    kept = [row for row in existing_rows if not row or (len(row) > pub_idx and row[pub_idx] != pub_iso)]

    new_rows_out: list[list[str]] = []
    for row in rows:
        s_row = ["" if v is None else str(v) for v in row]
        h = hashlib.sha256(("|".join(s_row) + "|" + pub_iso).encode("utf-8")).hexdigest()[:16]
        new_rows_out.append(s_row + [pub_iso, h])

    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(out_header)
        w.writerows(kept)
        w.writerows(new_rows_out)
    return len(new_rows_out)


# ---------------------------------------------------------------------------
# HTML table pages (distro, ownshares)
# ---------------------------------------------------------------------------

def _iso_from_impl_date(s: str) -> str:
    # '2026/06/26' -> '2026-06-26'; fallback: return as-is
    s = _norm_ws(s)
    m = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})$", s)
    if not m:
        return s
    return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"


def _pick_target_table(tables: list[list[tuple[str, list[str]]]], expected_headers: list[str]
                      ) -> tuple[list[tuple[str, list[str]]] | None, list[str] | None]:
    """Choose the table whose flattened header contains ALL expected headers.

    ヘッダの ``<br>`` はレイアウト由来なので、比較時のみ全空白を除去して照合する
    (例: ``買付申込数量<br>の限度`` = ``買付申込数量の限度``)。
    Returns (table, actual_flattened_headers) or (None, None).
    """
    def _strip_all_ws(s: str) -> str:
        return re.sub(r"\s+", "", s)
    exp_norm = {_strip_all_ws(h) for h in expected_headers}
    for tbl in tables:
        hdrs = _flatten_header_cells(tbl)
        hdrs_norm = {_strip_all_ws(h) for h in hdrs}
        if exp_norm.issubset(hdrs_norm):
            return tbl, hdrs
    return None, None


def _fetch_page_distro(cfg_page: dict[str, Any], http: Http, storage: dict[str, str],
                       now_tz: datetime, root: Path, log: logging.Logger) -> PageResult:
    page = "offauction_distribution"
    res = PageResult(page=page)
    raw_dir_root = root / storage["raw_root"] / page
    parsed_csv = root / storage["parsed_root"] / f"{page}.csv"
    manifest = raw_dir_root / storage["manifest_filename"]
    capture_iso = now_tz.strftime("%Y-%m-%d")
    raw_dir = raw_dir_root / capture_iso
    raw_dir.mkdir(parents=True, exist_ok=True)
    parsed_csv.parent.mkdir(parents=True, exist_ok=True)

    url = cfg_page["url"]
    try:
        html_bytes = http.get(url, log=log)
    except Exception as e:
        log.error("distro fetch failed: %s", e)
        res.fetch_errors.append(f"{url}: {e}")
        _append_manifest(manifest, ManifestEntry(
            captured_at=now_tz.isoformat(timespec="seconds"),
            page=page, date_key=capture_iso, source_url=url,
            local_path="", sha256="", bytes=0, rows_parsed=None,
            status="fetch_error", detail=str(e),
        ))
        return res

    local_path = raw_dir / "index.html"
    local_path.write_bytes(html_bytes)
    sha = _sha256(html_bytes)

    tables = _extract_tables(html_bytes)
    tbl, hdrs = _pick_target_table(tables, cfg_page["expected_headers"])
    if tbl is None:
        detail = f"header mismatch: expected all of {cfg_page['expected_headers']!r} in a single <table>"
        log.error("distro schema mismatch: %s", detail)
        res.schema_mismatches.append(detail)
        _append_manifest(manifest, ManifestEntry(
            captured_at=now_tz.isoformat(timespec="seconds"),
            page=page, date_key=capture_iso, source_url=url,
            local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(html_bytes),
            rows_parsed=None, status="schema_mismatch", detail=detail,
        ))
        return res

    # extract paired td rows.
    td_rows = [cells for kind, cells in tbl if kind == "td"]
    a_n = int(cfg_page["tr_a_cell_count"])
    b_n = int(cfg_page["tr_b_cell_count"])
    records: list[dict[str, str]] = []
    i = 0
    pair_errors: list[str] = []
    while i < len(td_rows):
        tr_a = td_rows[i]
        if len(tr_a) != a_n:
            pair_errors.append(f"row {i}: td count {len(tr_a)} != tr_a_cell_count {a_n}")
            i += 1
            continue
        if i + 1 >= len(td_rows):
            pair_errors.append(f"row {i}: missing pair second row")
            break
        tr_b = td_rows[i + 1]
        if len(tr_b) != b_n:
            pair_errors.append(f"row {i+1}: td count {len(tr_b)} != tr_b_cell_count {b_n}")
            i += 2
            continue
        impl_raw = tr_a[0]
        issue_raw = tr_a[1]
        prev_close = tr_a[2]
        offered_shares = tr_a[3]
        buy_limit = tr_a[4]
        purpose = tr_a[5]
        other_exchanges = tr_a[6]
        distro_price = tr_b[0]
        executed_shares = tr_b[1]
        name, code = _split_name_code(issue_raw)
        records.append({
            "implementation_date": _iso_from_impl_date(impl_raw),
            "implementation_date_raw": impl_raw,
            "issue_name": name,
            "issue_code": code,
            "issue_name_code_raw": issue_raw,
            "prev_close_yen": _strip_yen_suffix_int(prev_close),
            "prev_close_raw": prev_close,
            "distribution_price_yen": _strip_yen_suffix_int(distro_price),
            "distribution_price_raw": distro_price,
            "offered_shares": _strip_yen_suffix_int(offered_shares),
            "offered_shares_raw": offered_shares,
            "executed_shares": _strip_yen_suffix_int(executed_shares),
            "executed_shares_raw": executed_shares,
            "buy_limit_per_customer_raw": buy_limit,
            "purpose_raw": purpose,
            "other_exchanges_raw": other_exchanges,
        })
        i += 2

    if pair_errors:
        detail = "; ".join(pair_errors[:5]) + (" ..." if len(pair_errors) > 5 else "")
        log.error("distro row-pair mismatch: %s", detail)
        res.schema_mismatches.append(detail)
        _append_manifest(manifest, ManifestEntry(
            captured_at=now_tz.isoformat(timespec="seconds"),
            page=page, date_key=capture_iso, source_url=url,
            local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(html_bytes),
            rows_parsed=None, status="schema_mismatch", detail=detail,
        ))
        return res

    added = _append_records_by_natural_key(parsed_csv, records, cfg_page["natural_key"])
    res.new_rows += added
    res.ok_files += 1
    _append_manifest(manifest, ManifestEntry(
        captured_at=now_tz.isoformat(timespec="seconds"),
        page=page, date_key=capture_iso, source_url=url,
        local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(html_bytes),
        rows_parsed=len(records), status="ok", detail=f"new_rows={added}",
    ))
    return res


def _fetch_page_ownshares(cfg_page: dict[str, Any], http: Http, storage: dict[str, str],
                         now_tz: datetime, root: Path, log: logging.Logger) -> PageResult:
    page = "buyback_ownshares"
    res = PageResult(page=page)
    raw_dir_root = root / storage["raw_root"] / page
    parsed_csv = root / storage["parsed_root"] / f"{page}.csv"
    manifest = raw_dir_root / storage["manifest_filename"]
    capture_iso = now_tz.strftime("%Y-%m-%d")
    raw_dir = raw_dir_root / capture_iso
    raw_dir.mkdir(parents=True, exist_ok=True)
    parsed_csv.parent.mkdir(parents=True, exist_ok=True)

    url = cfg_page["url"]
    try:
        html_bytes = http.get(url, log=log)
    except Exception as e:
        log.error("ownshares fetch failed: %s", e)
        res.fetch_errors.append(f"{url}: {e}")
        _append_manifest(manifest, ManifestEntry(
            captured_at=now_tz.isoformat(timespec="seconds"),
            page=page, date_key=capture_iso, source_url=url,
            local_path="", sha256="", bytes=0, rows_parsed=None,
            status="fetch_error", detail=str(e),
        ))
        return res

    local_path = raw_dir / "index.html"
    local_path.write_bytes(html_bytes)
    sha = _sha256(html_bytes)

    tables = _extract_tables(html_bytes)
    tbl, hdrs = _pick_target_table(tables, cfg_page["expected_headers"])
    if tbl is None:
        detail = f"header mismatch: expected {cfg_page['expected_headers']!r}"
        log.error("ownshares schema mismatch: %s", detail)
        res.schema_mismatches.append(detail)
        _append_manifest(manifest, ManifestEntry(
            captured_at=now_tz.isoformat(timespec="seconds"),
            page=page, date_key=capture_iso, source_url=url,
            local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(html_bytes),
            rows_parsed=None, status="schema_mismatch", detail=detail,
        ))
        return res

    records: list[dict[str, str]] = []
    for kind, cells in tbl:
        if kind != "td":
            continue
        if len(cells) != 5:
            detail = f"td count {len(cells)} != 5"
            log.error("ownshares row mismatch: %s", detail)
            res.schema_mismatches.append(detail)
            _append_manifest(manifest, ManifestEntry(
                captured_at=now_tz.isoformat(timespec="seconds"),
                page=page, date_key=capture_iso, source_url=url,
                local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(html_bytes),
                rows_parsed=None, status="schema_mismatch", detail=detail,
            ))
            return res
        impl_raw, issue_raw, price_raw, buy_raw, exec_raw = cells
        name, code = _split_name_code(issue_raw)
        records.append({
            "implementation_date": _iso_from_impl_date(impl_raw),
            "implementation_date_raw": impl_raw,
            "issue_name": name,
            "issue_code": code,
            "issue_name_code_raw": issue_raw,
            "price_yen": _strip_yen_suffix_int(price_raw),
            "price_raw": price_raw,
            "buy_shares": _strip_yen_suffix_int(buy_raw),
            "buy_shares_raw": buy_raw,
            "executed_shares": _strip_yen_suffix_int(exec_raw),
            "executed_shares_raw": exec_raw,
        })

    added = _append_records_by_natural_key(parsed_csv, records, cfg_page["natural_key"])
    res.new_rows += added
    res.ok_files += 1
    _append_manifest(manifest, ManifestEntry(
        captured_at=now_tz.isoformat(timespec="seconds"),
        page=page, date_key=capture_iso, source_url=url,
        local_path=str(local_path.relative_to(root)), sha256=sha, bytes=len(html_bytes),
        rows_parsed=len(records), status="ok", detail=f"new_rows={added}",
    ))
    return res


def _append_records_by_natural_key(csv_path: Path, records: list[dict[str, str]],
                                   natural_key: list[str]) -> int:
    """Append records not yet present (by natural_key tuple). Returns new-row count."""
    if not records:
        return 0
    all_columns = list(records[0].keys())
    # ensure natural_key columns exist
    for k in natural_key:
        if k not in all_columns:
            raise ValueError(f"natural_key column {k!r} not in record columns {all_columns!r}")

    seen_keys: set[tuple[str, ...]] = set()
    existing_header: list[str] | None = None
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8", newline="") as f:
            r = csv.reader(f)
            existing_header = next(r, None)
            if existing_header is not None:
                # column set must be a superset (allow order tolerance? no: enforce exact for now).
                if existing_header != all_columns:
                    raise ValueError(
                        f"CSV header mismatch on {csv_path}: file has {existing_header!r} "
                        f"vs expected {all_columns!r}"
                    )
                idx = [existing_header.index(k) for k in natural_key]
                for row in r:
                    if not row:
                        continue
                    seen_keys.add(tuple(row[i] for i in idx))

    to_write: list[list[str]] = []
    for rec in records:
        key = tuple(rec[k] for k in natural_key)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        to_write.append([rec[c] for c in all_columns])

    write_header = existing_header is None
    with csv_path.open("a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if write_header:
            w.writerow(all_columns)
        w.writerows(to_write)
    return len(to_write)


# ---------------------------------------------------------------------------
# gaps report
# ---------------------------------------------------------------------------

def _append_gaps_report(path: Path, when: datetime, results: list[PageResult]) -> None:
    lines: list[str] = []
    for r in results:
        for sm in r.schema_mismatches:
            lines.append(f"- {when.isoformat(timespec='seconds')} | {r.page} | schema_mismatch | {sm}")
        for fe in r.fetch_errors:
            lines.append(f"- {when.isoformat(timespec='seconds')} | {r.page} | fetch_error | {fe}")
    if not lines:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    header_needed = not path.exists()
    with path.open("a", encoding="utf-8") as f:
        if header_needed:
            f.write("# unwind-tape gaps report\n\n")
            f.write("`ts | page | kind | detail` — schema変化 / 取得失敗 / 欠損の追記ログ。\n")
            f.write("欠損は絶対に埋めない。ここに列挙し、原因を追跡してから raw を再取得すること。\n\n")
        for l in lines:
            f.write(l + "\n")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def _configure_logging(log_dir: Path, when: datetime) -> logging.Logger:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"jpx_fetch_{when.strftime('%Y-%m-%d')}.log"
    logger = logging.getLogger("unwind_tape.jpx_fetch")
    logger.setLevel(logging.INFO)
    # avoid duplicate handlers on repeated invocations in same process
    logger.handlers.clear()
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    logger.addHandler(fh)
    sh = logging.StreamHandler(sys.stderr)
    sh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    logger.addHandler(sh)
    return logger


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Fetch JPX 立会外取引 information (unwind-tape task A)")
    ap.add_argument("--config", type=Path,
                    default=Path(__file__).resolve().parent.parent / "configs" / "jpx_offauction.yaml")
    ap.add_argument("--root", type=Path, default=Path(__file__).resolve().parent.parent,
                    help="unwind-tape root (data/, configs/ live here)")
    ap.add_argument("--pages", nargs="*", default=None,
                    help="subset of page keys to fetch (default: all in config)")
    ap.add_argument("--max-stale-business-days", type=int, default=5,
                    help="freshness alarm threshold (営業日換算、祝日非考慮)")
    ap.add_argument("--skip-freshness-check", action="store_true",
                    help="鮮度アラームを無効化する(デバッグ用)")
    args = ap.parse_args(argv)

    cfg = _load_config(args.config)
    tz = ZoneInfo(cfg["runtime"]["timezone"])
    now = _now(tz)
    log = _configure_logging(args.root / cfg["storage"]["log_dir"], now)
    log.info("unwind-tape jpx fetch start (root=%s, config=%s, now=%s)", args.root, args.config, now.isoformat())

    http = Http(
        ua=cfg["http"]["user_agent"],
        timeout=float(cfg["http"]["timeout_sec"]),
        retries=int(cfg["http"]["retries"]),
        base=float(cfg["http"]["backoff_base_sec"]),
        factor=float(cfg["http"]["backoff_factor"]),
    )
    storage = cfg["storage"]

    handlers = {
        "xlsx_index": _fetch_page_tostnet,
        "html_table_rowspan2": _fetch_page_distro,
        "html_table_simple": _fetch_page_ownshares,
    }

    selected = args.pages or list(cfg["pages"].keys())
    results: list[PageResult] = []
    exit_code = 0
    for page_key in selected:
        pcfg = cfg["pages"].get(page_key)
        if pcfg is None:
            log.error("no config for page %r", page_key)
            exit_code = 2
            continue
        h = handlers.get(pcfg["kind"])
        if h is None:
            log.error("no handler for kind %r on page %r", pcfg["kind"], page_key)
            exit_code = 2
            continue
        try:
            r = h(pcfg, http, storage, now, args.root, log)
        except Exception:
            log.error("uncaught exception on %s:\n%s", page_key, traceback.format_exc())
            r = PageResult(page=page_key, fetch_errors=[f"uncaught: {traceback.format_exc().strip().splitlines()[-1]}"])
        results.append(r)
        log.info("page=%s ok_files=%d new_rows=%d dup=%d schema_mismatches=%d fetch_errors=%d",
                 r.page, r.ok_files, r.new_rows, r.duplicates_skipped,
                 len(r.schema_mismatches), len(r.fetch_errors))
        if r.schema_mismatches:
            exit_code = max(exit_code, 3)
        if r.fetch_errors:
            exit_code = max(exit_code, 4)

    _append_gaps_report(args.root / storage["gaps_report"], now, results)

    if not args.skip_freshness_check:
        stale = _check_freshness(args.root, storage, selected, now,
                                 args.max_stale_business_days, log)
        if stale:
            exit_code = max(exit_code, 5)

    log.info("unwind-tape jpx fetch done (exit=%d)", exit_code)
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
