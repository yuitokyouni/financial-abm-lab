#!/usr/bin/env python3
"""unwind-tape / task B step 4 — CSV → xlsx round-trip.

CSV (groups/legs/sources + metadata) から xlsx を再構築する。
以後の運用: CSV を編集 → build_tape.py で xlsx を regenerate。

出力: data/parsed/tape/policy_holding_sale_event_tape_regenerated.xlsx
      9 sheets (Dashboard / Event_Tape / Field_Dictionary / Lists / Baseline_Spec /
                Source_Log / README / Sampling_Frame / Changelog)

不変条件:
  - validator を先に呼び ERROR があれば ABORT (build しない)
  - Event_Tape の行順は legs.csv の登場順 (groups で issuer 情報を hydrate)
  - v0.3 の 69 列順を厳守
  - 数値 → int/float、空欄 → None (Excel の空セル)
  - date → date object、bool → True/False
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import yaml


# import the v0.3 canonical column order + group/formula sets from the migration script.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from migrate_xlsx_to_csv import (  # noqa: E402
    EVENT_TAPE_COLUMNS_V03, GROUP_COLUMNS, FORMULA_COLUMNS, BOOL_FIELDS,
    DATE_FIELDS, TIME_FIELDS, NUMBER_FIELDS,
)


# ---------------------------------------------------------------------------
# cell coercers (string → typed value)
# ---------------------------------------------------------------------------

def _to_bool(s: str):
    if not s:
        return None
    s = s.strip().upper()
    if s == "TRUE":
        return True
    if s == "FALSE":
        return False
    return s  # UNKNOWN etc.


def _to_date(s: str):
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if not m:
        return s
    return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))


def _to_time(s: str):
    if not s:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if not m:
        return s
    return dt.time(int(m.group(1)), int(m.group(2)))


def _to_number(s: str):
    if not s:
        return None
    try:
        v = float(s)
        if v.is_integer():
            return int(v)
        return v
    except ValueError:
        return s


def _coerce(field: str, s: str):
    if field in BOOL_FIELDS or field == "activist_pressure":
        return _to_bool(s)
    if field in DATE_FIELDS:
        return _to_date(s)
    if field in TIME_FIELDS:
        return _to_time(s)
    if field in NUMBER_FIELDS:
        return _to_number(s)
    return s if s else None


# ---------------------------------------------------------------------------
# CSV / YAML loaders
# ---------------------------------------------------------------------------

def _load_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        rows = list(r)
        return list(r.fieldnames or []), rows


def _load_yaml(path: Path):
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


# ---------------------------------------------------------------------------
# sheet builders
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")


def _write_header(ws, columns: list[str]) -> None:
    for j, c in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def build_event_tape(wb: Workbook, groups: list[dict], legs: list[dict],
                     lists: dict[str, list[str]]) -> None:
    ws = wb.create_sheet("Event_Tape")
    _write_header(ws, EVENT_TAPE_COLUMNS_V03)

    group_by_id = {g["event_group_id"]: g for g in groups}
    for i, leg in enumerate(legs, start=2):
        gid = leg["event_group_id"]
        grec = group_by_id.get(gid, {})
        for j, col in enumerate(EVENT_TAPE_COLUMNS_V03, start=1):
            # sources: use URL columns; source_id_* are derived, don't write to Event_Tape
            if col in GROUP_COLUMNS and col != "event_group_id":
                v = grec.get(col, "")
            else:
                v = leg.get(col, "")
            # formula columns: leave blank (Task C fills; v0.3 has them empty)
            if col in FORMULA_COLUMNS:
                ws.cell(row=i, column=j, value=None)
                continue
            ws.cell(row=i, column=j, value=_coerce(col, v))

    # freeze first 3 columns per v0.3 audit change C024
    ws.freeze_panes = "D2"

    # data validation: enum + bool. Use Lists sheet ranges (added below).
    # v0.3 change C023: enum columns get list validation.
    lists_ranges = {  # name -> (col_letter_in_Lists, count)
        "status": ("A", 6),
        "event_tier": ("B", 3),
        "event_role": ("C", 8),
        "sale_route": ("D", 9),
        "absorption_route": ("E", 8),
        "seller_type": ("F", 9),
        "quantity_basis": ("G", 5),
        "value_basis": ("H", 5),
        "ABM_candidate_flag": ("I", 3),
        "confidence_policy_holding": ("J", 4),
        "activist_pressure": ("K", 3),
        "bool": ("L", 2),
    }
    for col in EVENT_TAPE_COLUMNS_V03:
        list_key = None
        if col in {"status", "event_tier", "event_role", "sale_route", "absorption_route",
                   "seller_type", "quantity_basis", "value_basis", "ABM_candidate_flag",
                   "confidence_policy_holding", "activist_pressure"}:
            list_key = col
        elif col in BOOL_FIELDS:
            list_key = "bool"
        if list_key is None or list_key not in lists_ranges:
            continue
        col_letter_lists, count = lists_ranges[list_key]
        col_idx = EVENT_TAPE_COLUMNS_V03.index(col) + 1
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        formula = f"Lists!${col_letter_lists}$2:${col_letter_lists}${count + 1}"
        dv = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
        dv.error = f"{col}: Lists.{list_key} 外の値"
        dv.errorTitle = "enum"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}400")


def build_lists(wb: Workbook, lists: dict[str, list[str]]) -> None:
    ws = wb.create_sheet("Lists")
    keys = list(lists.keys())
    for j, k in enumerate(keys, start=1):
        cell = ws.cell(row=1, column=j, value=k)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        for i, v in enumerate(lists[k], start=2):
            ws.cell(row=i, column=j, value=v)


def build_field_dictionary(wb: Workbook, rows: list[dict], columns: list[str]) -> None:
    ws = wb.create_sheet("Field_Dictionary")
    _write_header(ws, columns)
    for i, r in enumerate(rows, start=2):
        for j, c in enumerate(columns, start=1):
            ws.cell(row=i, column=j, value=r.get(c) or None)


def build_baseline_spec(wb: Workbook, rows: list[dict], columns: list[str]) -> None:
    ws = wb.create_sheet("Baseline_Spec")
    _write_header(ws, columns)
    for i, r in enumerate(rows, start=2):
        for j, c in enumerate(columns, start=1):
            ws.cell(row=i, column=j, value=r.get(c) or None)


def build_source_log(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Source_Log")
    columns = ["source_id", "source_type", "source_url", "what_it_supports", "archived", "notes"]
    _write_header(ws, columns)
    for i, r in enumerate(rows, start=2):
        for j, c in enumerate(columns, start=1):
            v = r.get(c, "")
            if c == "archived":
                ws.cell(row=i, column=j, value=_to_bool(v))
            else:
                ws.cell(row=i, column=j, value=v or None)


def build_sampling_frame(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Sampling_Frame")
    for i, r in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=r.get("key") or None)
        ws.cell(row=i, column=2, value=r.get("value") or None)


def build_changelog(wb: Workbook, rows: list[dict], columns: list[str]) -> None:
    ws = wb.create_sheet("Changelog")
    _write_header(ws, columns)
    for i, r in enumerate(rows, start=2):
        for j, c in enumerate(columns, start=1):
            ws.cell(row=i, column=j, value=r.get(c) or None)


def build_readme(wb: Workbook, readme: dict) -> None:
    ws = wb.create_sheet("README")
    for i, (k, v) in enumerate(readme.items(), start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)


def build_dashboard(wb: Workbook, groups: list[dict], legs: list[dict], version_tag: str) -> None:
    ws = wb.create_sheet("Dashboard", 0)  # first sheet
    ws.cell(row=1, column=1, value=f"Policy Holding Sale Event Tape {version_tag} (regenerated)").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="counts (formulas over Event_Tape)").font = HEADER_FONT

    # Restored formulas per v0.3 change C021 — counts over Event_Tape rows 2:400.
    # Positions of columns in Event_Tape.
    def col_letter(name: str) -> str:
        return openpyxl.utils.get_column_letter(EVENT_TAPE_COLUMNS_V03.index(name) + 1)

    et = "Event_Tape"
    formulas = [
        ("total legs (non-empty status)",       f"=COUNTA({et}!{col_letter('status')}2:{col_letter('status')}400)"),
        ("Tier1_confirmed legs",                f'=COUNTIF({et}!{col_letter("event_tier")}2:{col_letter("event_tier")}400,"Tier1_confirmed")'),
        ("Tier2_candidate legs",                f'=COUNTIF({et}!{col_letter("event_tier")}2:{col_letter("event_tier")}400,"Tier2_candidate")'),
        ("Tier3_overhang legs",                 f'=COUNTIF({et}!{col_letter("event_tier")}2:{col_letter("event_tier")}400,"Tier3_overhang")'),
        ("status=seeded",                       f'=COUNTIF({et}!{col_letter("status")}2:{col_letter("status")}400,"seeded")'),
        ("status=candidate",                    f'=COUNTIF({et}!{col_letter("status")}2:{col_letter("status")}400,"candidate")'),
        ("status=needs_source_check",           f'=COUNTIF({et}!{col_letter("status")}2:{col_letter("status")}400,"needs_source_check")'),
        ("status=needs_price_data",             f'=COUNTIF({et}!{col_letter("status")}2:{col_letter("status")}400,"needs_price_data")'),
        ("status=complete",                     f'=COUNTIF({et}!{col_letter("status")}2:{col_letter("status")}400,"complete")'),
        ("ABM_candidate_flag=Yes",              f'=COUNTIF({et}!{col_letter("ABM_candidate_flag")}2:{col_letter("ABM_candidate_flag")}400,"Yes")'),
        ("ABM_candidate_flag=No",               f'=COUNTIF({et}!{col_letter("ABM_candidate_flag")}2:{col_letter("ABM_candidate_flag")}400,"No")'),
        ("ABM_candidate_flag=Maybe",            f'=COUNTIF({et}!{col_letter("ABM_candidate_flag")}2:{col_letter("ABM_candidate_flag")}400,"Maybe")'),
        ("confidence=A_explicit_policy_holding", f'=COUNTIF({et}!{col_letter("confidence_policy_holding")}2:{col_letter("confidence_policy_holding")}400,"A_explicit_policy_holding")'),
        ("confidence=B_strong_inference",       f'=COUNTIF({et}!{col_letter("confidence_policy_holding")}2:{col_letter("confidence_policy_holding")}400,"B_strong_inference")'),
        ("confidence=C_possible_only",          f'=COUNTIF({et}!{col_letter("confidence_policy_holding")}2:{col_letter("confidence_policy_holding")}400,"C_possible_only")'),
        ("confidence=D_not_policy_holding",     f'=COUNTIF({et}!{col_letter("confidence_policy_holding")}2:{col_letter("confidence_policy_holding")}400,"D_not_policy_holding")'),
    ]
    for i, (label, formula) in enumerate(formulas, start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=formula)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tape-dir", type=Path,
                    default=Path(__file__).resolve().parent.parent / "data" / "parsed" / "tape")
    ap.add_argument("--out", type=Path,
                    default=Path(__file__).resolve().parent.parent / "data" / "parsed" / "tape"
                    / "policy_holding_sale_event_tape_regenerated.xlsx")
    ap.add_argument("--version-tag", default="v0.3")
    ap.add_argument("--skip-validate", action="store_true",
                    help="do NOT run validator first (default runs it)")
    args = ap.parse_args(argv)

    if not args.skip_validate:
        import subprocess
        r = subprocess.run(
            [sys.executable, str(Path(__file__).parent / "validate_tape.py"),
             "--tape-dir", str(args.tape_dir)],
            capture_output=True, text=True,
        )
        print(r.stdout, end="")
        print(r.stderr, end="", file=sys.stderr)
        if r.returncode == 4:
            print("build_tape: validator reported ERRORs, aborting build.", file=sys.stderr)
            return r.returncode
        # WARN (rc=3) is OK — proceed.

    _, groups = _load_csv(args.tape_dir / "groups.csv")
    _, legs = _load_csv(args.tape_dir / "legs.csv")
    _, sources = _load_csv(args.tape_dir / "sources.csv")
    lists = _load_yaml(args.tape_dir / "lists.yaml")
    fd_cols, field_dict = _load_csv(args.tape_dir / "field_dictionary.csv")
    _, sampling = _load_csv(args.tape_dir / "sampling_frame.csv")
    bs_cols, baseline = _load_csv(args.tape_dir / "baseline_spec.csv")
    cl_cols, changelog = _load_csv(args.tape_dir / "changelog.csv")
    readme = _load_yaml(args.tape_dir / "readme.yaml") or {}

    wb = Workbook()
    # remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # v0.3 order: Dashboard, Event_Tape, Field_Dictionary, Lists, Baseline_Spec,
    #             Source_Log, README, Sampling_Frame, Changelog
    build_dashboard(wb, groups, legs, args.version_tag)
    build_event_tape(wb, groups, legs, lists)
    build_field_dictionary(wb, field_dict, fd_cols)
    build_lists(wb, lists)
    build_baseline_spec(wb, baseline, bs_cols)
    build_source_log(wb, sources)
    build_readme(wb, readme)
    build_sampling_frame(wb, sampling)
    build_changelog(wb, changelog, cl_cols)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
